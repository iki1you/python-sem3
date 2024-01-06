import re
import numpy as np
import pandas as pd
import matplotlib
from matplotlib import pyplot as plt
from matplotlib.axes import Axes
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def get_medium(x):
    if x['salary_from'].equals(x['salary_to']):
        return x['salary_from']
    return (x['salary_from'] + x['salary_to']) / 2


def get_year_vacancy(s):
    for j in re.findall(r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\+\d{4}', s):
        s = s.replace(j, j[0:4])
    return int(s)


def get_vac_by_years(data):
    d = data.groupby(['published_at']).agg({
        'medium_salary': 'mean',
        'count': 'count',
    }).assign(medium_salary=lambda x: np.round(x['medium_salary']))
    return d


def get_vac_by_city(data):
    k = data.shape[0]
    d = (data.groupby(['area_name']).agg({
        'medium_salary': 'mean',
        'count': 'count',
    })
         .assign(count=lambda x: round(x['count'] / k * 100, 2))
         .query('count >= 1')
         .assign(medium_salary=lambda x: np.round(x['medium_salary'])))
    return (d.sort_values(['medium_salary', 'area_name'], ascending=(False, True))[:10]['medium_salary'],
            d.sort_values(['count', 'area_name'], ascending=(False, True))[:10]['count'])


def create_report():
    csv = 'vacancies.csv'

    vacancies = parse_csv(csv)

    d = get_vac_by_years(vacancies)
    wb = Workbook()
    wb.create_sheet(title='Статистика по годам', index=0)  # Создаем лист
    wb.create_sheet(title='Статистика по городам', index=1)
    sheetnames = wb.sheetnames  # Получение списка листов книги
    wb.remove(wb[sheetnames[2]])
    sheet = wb[sheetnames[0]]
    sheet.append(['Год', 'Средняя зарплата', 'Количество вакансий'])
    for i, r in enumerate(dataframe_to_rows(d, index=True, header=True)):
        if i >= 2:
            sheet.append(r)

    salary_vac, count_vac = get_vac_by_city(vacancies)
    sheet = wb[sheetnames[1]]
    sheet.append(['Город', 'Уровень зарплат'])
    for i, r in enumerate(dataframe_to_rows(pd.DataFrame(salary_vac), index=True, header=True)):
        if i >= 2:
            sheet.append(r)

    sheet.append(['Город', 'Доля вакансий, %'])
    for i, r in enumerate(dataframe_to_rows(pd.DataFrame(count_vac), index=True, header=True)):
        if i >= 2:
            sheet.append(r)
    sheet.move_range("A12:B22", rows=-11, cols=3)
    wb.save('student_works/report.xlsx')


def parse_csv(csv):
    names = ['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at']
    vacancies = (pd.read_csv(csv, names=names)
                 .assign(salary_from=lambda x: x['salary_from'].fillna(x['salary_to']))
                 .assign(salary_to=lambda x: x['salary_to'].fillna(x['salary_from']))
                 .assign(published_at=lambda x: x['published_at'].apply(get_year_vacancy))
                 .assign(count=0)
                 .assign(medium_salary=get_medium))
    return vacancies


def create_plot():

    csv = 'vacancies.csv'
    vac = input()
    vacancies = parse_csv(csv)
    fig, sub = plt.subplots(2, 2, figsize=(14, 9))
    vac_by_years = get_vac_by_years(vacancies)['medium_salary'].to_dict()
    vac_by_years_filtered = get_vac_by_years(vacancies[vacancies['name']
                                             .str.contains(vac, na=False, case=False)])['medium_salary'].to_dict()

    vac_by_years = dict([(key, value) for key, value in vac_by_years.items() if key in vac_by_years_filtered])

    ax: Axes = sub[0, 0]
    xlable = range(min(vac_by_years_filtered.keys()), max(vac_by_years_filtered.keys()) + 1)

    x = np.arange(len(xlable))

    y = [i for i in vac_by_years.values()]
    y2 = [i for i in vac_by_years_filtered.values()]
    width = 0.3
    ax.bar(x - width/2, y, width, label='средняя з/п')
    ax.bar(x + width/2, y2, width, label=f'з/п {vac}')
    ax.set_title('Уровень зарплат по годам', fontsize=8)
    ax.set_xticks(x)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=90, ha='right')
    ax.set_xticklabels(xlable)
    ax.tick_params(labelsize=8)
    ax.legend(fontsize=8)

    set_font_size(ax, 8)


    vac_by_years = get_vac_by_years(vacancies)['count'].to_dict()
    vac_by_years_filtered = get_vac_by_years(vacancies[vacancies['name']
                                             .str.contains(vac, na=False, case=False)])['count'].to_dict()
    vac_by_years = dict([(key, value) for key, value in vac_by_years.items() if key in vac_by_years_filtered])
    width = 0.3
    ax = sub[0, 1]
    xlable = range(min(vac_by_years_filtered.keys()), max(vac_by_years_filtered.keys()) + 1)
    x = np.arange(len(xlable))
    y = [i for i in vac_by_years.values()]
    y2 = [i for i in vac_by_years_filtered.values()]
    ax.bar(x - width / 2, y, width, label='Количество вакансий')
    ax.bar(x + width / 2, y2, width, label=f'Количество вакансий {vac}')

    ax.set_title('Количество вакансий по годам')
    ax.set_xticks(x)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=90, ha='right')
    ax.set_xticklabels(xlable)
    ax.legend()

    level_by_cities, count_by_cities = get_vac_by_city(vacancies)
    count_by_cities = count_by_cities.to_dict()

    width = 0.3
    ax = sub[1, 0]
    x = [i for i in reversed(level_by_cities.to_dict().keys())]
    y = [i for i in reversed(level_by_cities.to_dict().values())]
    ylable = range(0, int(max(y)) + 19999, 20000)
    ax.barh(x, y, width)
    ax.set_xticks(ylable)
    ax.set_title('Уровень зарплат по городам')

    ax.set_xticklabels(ax.get_xticklabels(), rotation=90)
    plt.show()
    return sub


def set_font_size(ax, size):
    for label in (ax.get_xticklabels() + ax.get_yticklabels()):
        label.set_fontsize(size)
    dx = 0.06
    dy = 0
    fig = plt.figure()
    offset = matplotlib.transforms.ScaledTranslation(dx, dy, fig.dpi_scale_trans)
    for label in ax.xaxis.get_majorticklabels():
        label.set_transform(label.get_transform() + offset)


create_plot()