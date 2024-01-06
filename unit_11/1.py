import re

import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def get_medium(x):
    if x['salary_from'].equals(x['salary_to']):
        return x['salary_from']
    return (x['salary_from'] + x['salary_to']) / 2


def get_year_vacancy(s):
    for j in re.findall(r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\+\d{4}', s):
        s = s.replace(j, j[0:4])
    return int(s)


def get_vac_by_years(data):
    d = data.groupby(['published_at']).agg({
        'medium_salary': 'mean',
        'count': 'count',
    }).assign(medium_salary=lambda x: np.round(x['medium_salary']))
    return d


def get_vac_by_city(data):
    k = data.shape[0]
    d = (data.groupby(['area_name']).agg({
        'medium_salary': 'mean',
        'count': 'count',
    })
         .assign(count=lambda x: round(x['count'] / k * 100, 2))
         .query('count >= 1')
         .assign(medium_salary=lambda x: np.round(x['medium_salary'])))
    return (d.sort_values(['medium_salary', 'area_name'], ascending=(False, True))[:10]['medium_salary'],
            d.sort_values(['count', 'area_name'], ascending=(False, True))[:10]['count'])


def create_report():
    csv = 'vacancies.csv'

    names = ['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at']
    vacancies = (pd.read_csv(csv, names=names)
                 .assign(salary_from=lambda x: x['salary_from'].fillna(x['salary_to']))
                 .assign(salary_to=lambda x: x['salary_to'].fillna(x['salary_from']))
                 .assign(published_at=lambda x: x['published_at'].apply(get_year_vacancy))
                 .assign(count=0)
                 .assign(medium_salary=get_medium))

    wb = Workbook()
    wb.create_sheet(title='Статистика по годам', index=0)
    wb.create_sheet(title='Статистика по городам', index=1)

    data = get_vac_by_years(vacancies)
    sheet = wb[wb.sheetnames[0]]
    append_table(sheet, data, ['Год', 'Средняя зарплата', 'Количество вакансий'])

    salary_vac, count_vac = get_vac_by_city(vacancies)
    sheet = wb[wb.sheetnames[1]]
    append_table(sheet, pd.DataFrame(salary_vac), ['Город', 'Уровень зарплат'])
    append_table(sheet, pd.DataFrame(count_vac), ['Город', 'Доля вакансий, %'])

    sheet.move_range("A12:B22", rows=-11, cols=3)
    wb.remove(wb[wb.sheetnames[2]])
    wb.save('student_works/report.xlsx')


def append_table(sheet, data, columns):
    sheet.append(columns)
    for i, r in enumerate(dataframe_to_rows(data, index=True, header=True)):
        if i >= 2:
            sheet.append(r)


create_report()